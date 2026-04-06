"""
Project: Data RAG System
Module: Document ETL Pipeline
Description: 
    - 하이브리드 문서(PDF, DOCX, PPTX, XLSX)의 구조적 추출 및 마크다운 변환
    - 재귀적 디렉토리 스캔 및 계층 구조 보존 적재
    - 다국어(KO, EN) OCR 및 시각적 레이아웃 분석 지원
    - Memory & VRAM 최적화 적용 (Thread Pool & Worker Limit)
Author: H4RU
Date: 2026-03-17
"""

import os
import time
import logging
from abc import ABC, abstractmethod
from pathlib import Path
from dataclasses import dataclass, field
from typing import List, Set, Optional, Iterator
# 메모리 누수 방지를 위해 multiprocessing Pool 사용
import multiprocessing
from tqdm import tqdm

from docling.document_converter import DocumentConverter, PdfFormatOption
from docling.datamodel.pipeline_options import PdfPipelineOptions, EasyOcrOptions
from docling.datamodel.base_models import InputFormat, ConversionStatus

import traceback
import zipfile
import openpyxl
import xml.etree.ElementTree as ET

# --- openpyxl 몽키패치 ---
# 엑셀 파일 내 잘못된 '인쇄 제목(Print_Titles)' 포맷으로 인한 openpyxl 크래시 방지
from openpyxl.worksheet.print_settings import PrintTitles
_orig_print_titles_from_string = PrintTitles.from_string
@classmethod
def _safe_print_titles_from_string(cls, value):
    try:
        return _orig_print_titles_from_string(value)
    except ValueError:
        return cls()
PrintTitles.from_string = _safe_print_titles_from_string
# -------------------------

# 전역 로거 세팅
logger = logging.getLogger("RAG_ETL")
logger.setLevel(logging.INFO)
fmt = logging.Formatter('[%(levelname)s] %(asctime)s - %(message)s', '%H:%M:%S')
ch = logging.StreamHandler()
ch.setFormatter(fmt)
logger.addHandler(ch)

class DocProcessingError(Exception):
    """문서 변환 실패 예외"""
    pass

@dataclass(frozen=True)
class ETLConfig:
    # 파이프라인 불변 설정
    input_dir: Path
    output_dir: Path
    # VRAM/RAM OOM 방지
    max_workers: int = 4
    target_exts: Set[str] = field(default_factory=lambda: {'.pdf', '.docx', '.pptx', '.xlsx', '.hwpx'})
    skip_exts: Set[str] = field(default_factory=lambda: {'.png', '.jpg', '.jpeg', '.gif', '.bmp', '.webp'})
    allowed_formats: List[InputFormat] = field(
        default_factory=lambda: [InputFormat.PDF, InputFormat.DOCX, InputFormat.PPTX, InputFormat.XLSX]
    )

class ArchivePreprocessor:
    """ZIP 파일 탐색 및 자동 압축 해제"""
    def __init__(self, target_dir: Path):
        self.target_dir = target_dir

    def run(self) -> None:
        zip_files = list(self.target_dir.rglob('*.zip'))
        if not zip_files:
            return
            
        logger.info(f"ZIP Archive 발견: {len(zip_files)}ea. Extract 시작")
        
        for zf in zip_files:
            if not zf.is_file():
                continue
                
            extract_dir = zf.with_suffix('') 
            
            if extract_dir.exists():
                continue
                
            try:
                with zipfile.ZipFile(zf, 'r') as zip_ref:
                    zip_ref.extractall(extract_dir)
                logger.debug(f"Unzip 성공: {zf.name}")
            except zipfile.BadZipFile:
                logger.error(f"ZIP file 손상: {zf.name}")
            except Exception as e:
                logger.error(f"Unzip 에러 [{zf.name}]: {e}")

class IFileScanner(ABC):
    """스캐너 인터페이스"""
    @abstractmethod
    def get_targets(self) -> Iterator[Path]:
        pass

class RecursiveScanner(IFileScanner):
    def __init__(self, config: ETLConfig):
        self.cfg = config

    def get_targets(self) -> Iterator[Path]:
        if not self.cfg.input_dir.exists():
            raise FileNotFoundError(f"경로 없음: {self.cfg.input_dir}")
            
        for f in self.cfg.input_dir.rglob('*'):
            if not f.is_file():
                continue
                
            # OS 임시 파일 및 숨김 파일 스킵
            if f.name.startswith('~$') or f.name.startswith('.'):
                continue

            ext = f.suffix.lower()
            if ext in self.cfg.target_exts:
                yield f
            elif ext in self.cfg.skip_exts:
                pass

# 워커 프로세스별 전역 캐시
_worker_converter = None

def _init_worker():
    """워커 프로세스 초기화 시 엔진 로드 방지 (지연 로딩 활용)"""
    global _worker_converter
    _worker_converter = None
    
    import warnings
    warnings.filterwarnings("ignore", category=UserWarning)
    
    # docling 등 기타 라이브러리의 불필요한 로그 억제
    import logging
    for name in logging.root.manager.loggerDict:
        if name != "RAG_ETL":
            logging.getLogger(name).setLevel(logging.CRITICAL)

def _xlsx_to_markdown(fpath: Path) -> str:
    """openpyxl을 이용해 모든 시트를 마크다운으로 변환"""
    wb = openpyxl.load_workbook(fpath, data_only=True)
    parts = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # 비어있지 않은 행만 수집
        rows = []
        for row in ws.iter_rows(values_only=True):
            if all(v is None for v in row):
                continue
            rows.append(row)
        if not rows:
            continue

        parts.append(f"## {sheet_name}\n")

        # 모든 행 중 최대 열 수 산정 (헤더 행이 가장 짧을 수도 있음)
        max_cols = max(len(r) for r in rows)

        def cell_str(v) -> str:
            s = "" if v is None else str(v)
            return s.replace("|", "\\|").replace("\n", " ").replace("\r", "")

        header = [cell_str(v) for v in rows[0]]
        # 헤더 열 수가 max_cols보다 적으면 패딩
        while len(header) < max_cols:
            header.append("")

        parts.append("| " + " | ".join(header) + " |")
        parts.append("| " + " | ".join(["---"] * max_cols) + " |")

        for row in rows[1:]:
            cells = [cell_str(v) for v in row]
            while len(cells) < max_cols:
                cells.append("")
            parts.append("| " + " | ".join(cells) + " |")

        parts.append("")
    wb.close()
    return "\n".join(parts)


def _hwpx_to_markdown(fpath: Path) -> str:
    """HWPX(ZIP+XML) 파일의 모든 섹션에서 텍스트를 추출해 마크다운으로 변환"""
    # Hancom XML 네임스페이스
    HP_NS = "http://www.hancom.co.kr/hwpml/2012/paragraph"
    HP_T = f"{{{HP_NS}}}t"
    HP_P = f"{{{HP_NS}}}p"

    parts = []
    with zipfile.ZipFile(fpath, 'r') as zf:
        # Contents/section*.xml 파일만 정렬하여 순서대로 처리
        section_files = sorted(
            [n for n in zf.namelist() if n.startswith("Contents/section") and n.endswith(".xml")]
        )
        if not section_files:
            # 네임스페이스 없이 텍스트 태그를 직접 탐색 (fallback)
            section_files = sorted(
                [n for n in zf.namelist() if "section" in n and n.endswith(".xml")]
            )

        for sec_name in section_files:
            try:
                xml_bytes = zf.read(sec_name)
                root = ET.fromstring(xml_bytes)
            except ET.ParseError as e:
                logger.warning(f"HWPX XML 파싱 실패 [{fpath.name}/{sec_name}]: {e}")
                continue

            for para in root.iter(HP_P):
                line_parts = []
                for t_elem in para.iter(HP_T):
                    if t_elem.text:
                        line_parts.append(t_elem.text)
                line = "".join(line_parts).strip()
                if line:
                    parts.append(line)

    return "\n\n".join(parts)


def _process_route(args):
    """multiprocessing.Pool에서 호출하기 위한 최상위 경로 함수"""
    fpath, input_dir, output_dir, allowed_formats = args
    return DocumentETL._global_process_single(fpath, input_dir, output_dir, allowed_formats)

class DocumentETL:
    """문서 추출 및 변환 코어 엔진"""
    def __init__(self, config: ETLConfig, scanner: IFileScanner):
        self.cfg = config
        self.scanner = scanner
        # 메인 프로세스에서는 엔진을 로드하지 않음 (메모리 절약)
        self.cfg.output_dir.mkdir(parents=True, exist_ok=True)
        logger.info(f"ETL 파이프라인 로드 완료 (Workers: {self.cfg.max_workers})")

    @staticmethod
    def _build_engine_static(allowed_formats) -> DocumentConverter:
        # 초기화
        ocr_opts = EasyOcrOptions(lang=["ko", "en"])
        pdf_opts = PdfPipelineOptions()
        pdf_opts.do_ocr = True
        pdf_opts.ocr_options = ocr_opts
        pdf_opts.do_table_structure = True
        
        # 옵션 명시적 매핑 (Dictionary)
        return DocumentConverter(
            allowed_formats=allowed_formats,
            format_options={
                InputFormat.PDF: PdfFormatOption(pipeline_options=pdf_opts)
            }
        )

    @staticmethod
    def _global_process_single(fpath: Path, input_dir: Path, output_dir: Path, allowed_formats) -> Optional[Path]:
        global _worker_converter
        start_t = time.time()
        size_mb = fpath.stat().st_size / (1024 * 1024)

        try:
            # [기능 추가] 이어하기 체크 로직
            rel_path = fpath.relative_to(input_dir)
            out_path = output_dir / rel_path.with_suffix('.md')
            
            if out_path.exists():
                return out_path

            ext = fpath.suffix.lower()
            # XLSX는 openpyxl로 직접 변환 (모든 시트 처리)
            if ext == '.xlsx':
                logger.debug(f"처리 시작 (XLSX/openpyxl): {fpath.name}")
                md_text = _xlsx_to_markdown(fpath)
                if not md_text.strip():
                    logger.error(f"Parse failed [{fpath.name}]: XLSX 내용 없음")
                    return None
            # HWPX는 ZIP+XML 직접 파싱
            elif ext == '.hwpx':
                logger.debug(f"처리 시작 (HWPX): {fpath.name}")
                md_text = _hwpx_to_markdown(fpath)
                if not md_text.strip():
                    logger.error(f"Parse failed [{fpath.name}]: HWPX 내용 없음")
                    return None
            else:
                # 개별 워커 프로세스에서 엔진 최초 1회 로드
                if _worker_converter is None:
                    _worker_converter = DocumentETL._build_engine_static(allowed_formats)

                logger.debug(f"처리 시작: {fpath.name}")
                res = _worker_converter.convert(fpath, raises_on_error=False)

                if res.status not in (ConversionStatus.SUCCESS, ConversionStatus.PARTIAL_SUCCESS):
                    logger.error(f"Parse failed [{fpath.name}]: Document conversion failed with status {res.status.value}")
                    return None

                md_text = res.document.export_to_markdown()
            
            out_path.parent.mkdir(parents=True, exist_ok=True)
            out_path.write_text(md_text, encoding="utf-8")
            
            elapsed = time.time() - start_t
            speed = size_mb / elapsed if elapsed > 0 else 0
            
            logger.debug(f"파싱 성공: {fpath.name} ({elapsed:.2f}s, {speed:.2f} mb/s)")
            return out_path
            
        except Exception as e:
            # 에러 Traceback 생략 및 간결한 로깅
            logger.error(f"Parse failed [{fpath.name}]: {e}")
            return None

    def execute(self) -> None:
        targets = list(self.scanner.get_targets())
        if not targets:
            logger.warning("처리 대상 없음")
            return

        logger.info(f"배치 실행: 총 {len(targets)}건")
        start_t = time.time()
        success_cnt = 0

        # Memory Leak 방지를 위해 maxtasksperchild 설정 적용
        # 각 프로세스는 50개 파일 처리 후 자동 종료 및 재생성되어 메모리를 반환함
        task_args = [
            (f, self.cfg.input_dir, self.cfg.output_dir, self.cfg.allowed_formats) 
            for f in targets
        ]

        with multiprocessing.Pool(
            processes=self.cfg.max_workers,
            initializer=_init_worker,
            maxtasksperchild=50
        ) as pool:
            # imap_unordered를 사용하여 실시간 프로그레스 바 연동
            for result in tqdm(pool.imap_unordered(_process_route, task_args), total=len(targets), desc="ETL Progress"):
                if result:
                    success_cnt += 1

        total_t = time.time() - start_t
        logger.info(f"배치 완료: {success_cnt}/{len(targets)} 성공 (총 {total_t:.1f}s)")

if __name__ == "__main__":
    # 설정 및 의존성 주입 (Dependency Injection)
    cfg = ETLConfig(
        input_dir=Path("./files"),
        output_dir=Path("./processed_md")
    )

    preprocessor = ArchivePreprocessor(target_dir=cfg.input_dir)
    preprocessor.run()

    scanner = RecursiveScanner(config=cfg)
    etl = DocumentETL(config=cfg, scanner=scanner)
    
    etl.execute()